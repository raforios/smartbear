'''
    Decimal precision auditor for the biweekly cotizaciones xlsx.

    Why this exists:
        The official PDF from the Ministry publishes mineral prices with two
        decimals (e.g. WÓLFRAM 64205.75). When the operator manually copies
        them into the consolidated Excel, those decimals are sometimes lost —
        Excel stores them as ints and the ingest script ends up with rounded
        values in `t_mining_prices`.

        This auditor does NOT touch the Excel. It only inspects every Diario
        sheet and flags rows that *probably* lost precision so the operator
        knows which cells to re-key against the PDF.

    Two signals are reported:

    1. **Per-cell type check** — every numeric cell is classified as:
         * `int`         → value is a Python int (decimals impossible).
         * `clean_float` → float with non-zero fractional part (likely OK).
         * `whole_float` → float whose value is integer-equal
                           (e.g. 64206.0): could be legitimate OR a rounded
                           reading. Reported only when the column has at
                           least one `clean_float`, which proves that the
                           mineral *can* carry decimals in this period.

    2. **Promedio mismatch** — for every (mineral, sheet) pair we compare:
         * the value in the trailing `Promedio` row, and
         * the arithmetic mean of the daily rows we ingest.
        A mismatch is the strongest evidence that the daily rows were
        truncated, because the operator computed the average BEFORE
        rounding.

    Usage:
        python -m scripts.audit_decimals
        python -m scripts.audit_decimals --source /custom/path.xlsx --tolerance 0.01
'''
import argparse
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl


DEFAULT_SOURCE = (
    Path(__file__).resolve().parents[4] / 'data' / 'cotizaciones_mineras_bolivia.xlsx'
)

# Tolerance used when comparing `Promedio` vs. arithmetic mean: anything below
# this margin is considered a match (floating-point + Excel rounding artifacts).
DEFAULT_TOLERANCE = 0.01


def _normalize_mineral_label(raw: str) -> str:
    '''
        Strips the unit suffix and trailing whitespace from a column header.
    '''
    return str(raw).split('\n', maxsplit = 1)[0].strip()


def _classify(value) -> str:
    '''
        Returns one of {'int', 'whole_float', 'clean_float'} for a numeric cell.
    '''
    if isinstance(value, bool):
        return 'int'
    if isinstance(value, int):
        return 'int'
    if isinstance(value, float):
        return 'whole_float' if value == int(value) else 'clean_float'
    return 'int'


def _format_value(value) -> str:
    '''
        Pretty-print helper that keeps decimals visible.
    '''
    if isinstance(value, float):
        if value == int(value):
            return f'{value:.2f}'
        return f'{value:g}'
    return str(value)


def _is_promedio_cell(day_value) -> bool:
    '''
        True when the row label denotes the trailing average row.
    '''
    return isinstance(day_value, str) and 'promedio' in day_value.lower()


def _analyze_sheet(sheet) -> Tuple[
    List[Dict],   # mineral_columns analysis
    Dict[str, float],  # promedio_row by mineral
    Dict[str, List[float]]  # daily values by mineral
]:
    '''
        Walks one Diario sheet and returns the per-column type counters,
        the `Promedio` row content and the list of daily values.
    '''
    rows = list(sheet.iter_rows(values_only = False))
    if len(rows) < 4:
        return [], {}, {}

    header_cells = rows[2]
    mineral_cols: Dict[int, str] = {}
    for idx, cell in enumerate(header_cells):
        if idx == 0 or cell.value is None:
            continue
        label = _normalize_mineral_label(cell.value)
        if label:
            mineral_cols[idx] = label

    promedio_by_mineral: Dict[str, float] = {}
    daily_by_mineral: Dict[str, List[float]] = defaultdict(list)
    per_column = {idx: {
        'mineral': label,
        'int': 0,
        'whole_float': 0,
        'clean_float': 0,
        'whole_float_samples': []
    } for idx, label in mineral_cols.items()}

    for row in rows[3:]:
        day_cell = row[0]
        if day_cell.value is None:
            continue
        is_avg = _is_promedio_cell(day_cell.value)
        for idx, label in mineral_cols.items():
            if idx >= len(row):
                continue
            cell = row[idx]
            if cell.value is None or cell.value == '':
                continue
            if not isinstance(cell.value, (int, float)):
                continue
            value = float(cell.value)
            if is_avg:
                promedio_by_mineral[label] = value
                continue
            daily_by_mineral[label].append(value)
            kind = _classify(cell.value)
            per_column[idx][kind] += 1
            if kind == 'whole_float' and len(per_column[idx]['whole_float_samples']) < 4:
                per_column[idx]['whole_float_samples'].append(
                    (day_cell.value, cell.value)
                )

    return list(per_column.values()), promedio_by_mineral, daily_by_mineral


def _print_sheet_section(
    sheet_name: str,
    columns: List[Dict],
    promedio_by_mineral: Dict[str, float],
    daily_by_mineral: Dict[str, List[float]],
    tolerance: float
) -> int:
    '''
        Renders the per-sheet findings and returns the number of warnings raised.
    '''
    print(f'\n── {sheet_name} ───────────────────────')

    warnings, suspect_columns = _report_columns(columns, promedio_by_mineral)
    warnings += _report_average_crosscheck(
        promedio_by_mineral, daily_by_mineral, tolerance
    )

    if suspect_columns:
        print('  → re-revisar contra el PDF: ' + ', '.join(suspect_columns))
    return warnings


def _report_columns(
    columns: List[Dict],
    promedio_by_mineral: Dict[str, float]
) -> Tuple[int, List[str]]:
    '''
        Prints one line per mineral column and flags the suspicious ones.

        A column is suspicious when it mixes decimals with whole numbers — the
        sign that the operator rounded some days by hand — or when every day is
        a whole number yet the Promedio row carries decimals, which cannot
        happen if the days really were integers.

        Args:
            columns (List[Dict]): Per-column counters gathered from the sheet.
            promedio_by_mineral (Dict[str, float]): The sheet's Promedio row.

        Returns:
            Tuple[int, List[str]]: Warnings raised and the suspicious minerals.
    '''
    warnings = 0
    suspect_columns: List[str] = []

    for col in columns:
        mineral = col['mineral']
        total = col['int'] + col['whole_float'] + col['clean_float']
        if total == 0:
            continue

        has_decimals = col['clean_float'] > 0
        rounded_by_hand = has_decimals and col['int'] > 0
        avg_row = promedio_by_mineral.get(mineral)
        integer_days_with_decimal_average = (
            col['int'] == total and avg_row is not None and avg_row != int(avg_row)
        )

        marker = '   '
        if rounded_by_hand or integer_days_with_decimal_average:
            marker = ' ! '
            suspect_columns.append(mineral)
            warnings += 1

        print(
            f'{marker}{mineral:22} '
            f'int={col["int"]:>3} whole_float={col["whole_float"]:>3} '
            f'clean_float={col["clean_float"]:>3}'
        )
        if col['whole_float_samples'] and has_decimals:
            preview = ', '.join(f'día {d}={_format_value(v)}'
                                for d, v in col['whole_float_samples'])
            print(f'     · sospechosos: {preview}')

    return warnings, suspect_columns


def _report_average_crosscheck(
    promedio_by_mineral: Dict[str, float],
    daily_by_mineral: Dict[str, List[float]],
    tolerance: float
) -> int:
    '''
        Recomputes each Promedio from the daily values and reports the gap.

        The published average must match the mean of the days it summarises;
        a difference beyond the tolerance means the sheet was edited after the
        average was computed, or the days were truncated.

        Args:
            promedio_by_mineral (Dict[str, float]): The sheet's Promedio row.
            daily_by_mineral (Dict[str, List[float]]): Daily values per mineral.
            tolerance (float): Largest accepted difference.

        Returns:
            int: Warnings raised.
    '''
    if not promedio_by_mineral:
        return 0

    warnings = 0
    print('  Cruce con fila Promedio:')
    for mineral, promedio_value in sorted(promedio_by_mineral.items()):
        daily = daily_by_mineral.get(mineral, [])
        if not daily:
            continue
        computed = sum(daily) / len(daily)
        delta = abs(computed - promedio_value)
        mismatched = delta > tolerance
        if mismatched:
            warnings += 1
        print(
            f'{" ✗ " if mismatched else "   "}{mineral:22} '
            f'Promedio={_format_value(promedio_value):>14}  '
            f'mean(días)={computed:14.4f}  Δ={delta:10.4f}'
        )
    return warnings


def audit(source: Path, tolerance: float) -> int:
    '''
        Top-level orchestration. Returns the total number of warnings raised
        (used as the process exit code).
    '''
    if not source.exists():
        message = f'Source file not found: {source}'
        print(message, file = sys.stderr)
        return -1

    workbook = openpyxl.load_workbook(source, data_only = True)
    diary_sheets = [s for s in workbook.sheetnames if 'diario' in s.lower()]
    if not diary_sheets:
        print('No Diario sheets detected. Nothing to audit.')
        return 0

    print('=' * 70)
    print(f'  Decimal precision audit — {source.name}')
    print(f'  Tolerance for Promedio mismatch: ±{tolerance:g}')
    print('=' * 70)

    total_warnings = 0
    for sheet_name in diary_sheets:
        columns, promedio, daily = _analyze_sheet(workbook[sheet_name])
        total_warnings += _print_sheet_section(
            sheet_name, columns, promedio, daily, tolerance
        )

    print('\n' + '=' * 70)
    if total_warnings == 0:
        print('  ✓ No se detectaron pérdidas de precisión.')
    else:
        print(f'  ! {total_warnings} hallazgo(s) que requieren revisión.')
        print('    Leyenda:')
        print('      "!"  columna mezcla enteros con decimales → algunos días')
        print('           probablemente fueron redondeados al copiar del PDF.')
        print('      "✗"  el promedio del Excel NO coincide con el promedio')
        print('           aritmético de los días → casi seguro hay decimales')
        print('           originales perdidos en uno o más días.')
    print('=' * 70)
    return total_warnings


def main() -> int:
    '''
        CLI entry point.
    '''
    parser = argparse.ArgumentParser(
        description = 'Audit the cotizaciones xlsx for lost decimal precision.'
    )
    parser.add_argument('--source', type = Path, default = DEFAULT_SOURCE,
                        help = f'Excel path. Default: {DEFAULT_SOURCE}.')
    parser.add_argument('--tolerance', type = float, default = DEFAULT_TOLERANCE,
                        help = ('Maximum delta between the `Promedio` row and '
                                'the arithmetic mean of the days to consider '
                                'them equivalent. Default: 0.01.'))
    args = parser.parse_args()
    warnings = audit(args.source, args.tolerance)
    return 0 if warnings <= 0 else 1


if __name__ == '__main__':
    raise SystemExit(main())
