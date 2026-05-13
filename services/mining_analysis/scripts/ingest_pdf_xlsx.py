'''
    One-shot ingest of biweekly Diario sheets extracted from scanned PDFs.

    Input file layout (produced manually from the scanned monthly PDFs):
      - Sheet name pattern: `{Mes} Q{1|2} - Diario`
      - Header row 2 holds the period title with the month/year, e.g.
        `PRIMERA QUINCENA DE ABRIL 2026 (días 1 al 15)`.
      - Row 3 holds column headers, with mineral names embedding the unit:
        `ESTAÑO\\n(L.F.)`, `WÓLFRAM\\n(T.M.F.)`, etc.
      - Subsequent rows hold the day number in column A and prices per mineral.
      - A trailing `Promedio` row is ignored (we recompute it ourselves).

    Mineral name matching is case-insensitive and accent-insensitive so the
    sheet headers map correctly to the OFFICIAL_MINERALS catalog.

    Usage:
        python -m scripts.ingest_pdf_xlsx --source /path/to/file.xlsx [--yes]
'''
import argparse
import re
import sys
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

from models.mining_analysis import Mineral, MiningPrice
from services.db_connection import ENGINE, get_db_session
from services.logger_config import custom_logger as logger
from services.mining_analysis import (
    OFFICIAL_MINERALS,
    _normalize_name,
    ensure_official_minerals,
)


SPANISH_MONTHS = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'setiembre': 9, 'octubre': 10,
    'noviembre': 11, 'diciembre': 12,
}


def _parse_period_title(title: str) -> Optional[Tuple[int, int]]:
    '''
    Extracts (year, month) from titles like 'PRIMERA QUINCENA DE ABRIL 2026 (...)'.
    Returns None when the title does not match the expected pattern.
    '''
    if not isinstance(title, str):
        return None
    match = re.search(r'\b(' + '|'.join(SPANISH_MONTHS) + r')\b\s+(\d{4})',
                      title, flags = re.IGNORECASE)
    if not match:
        return None
    return int(match.group(2)), SPANISH_MONTHS[match.group(1).lower()]


def _extract_mineral_columns(header_row: Tuple) -> Dict[int, str]:
    '''
    Returns {column_index: normalized_mineral_name} for each header cell that
    matches an entry in OFFICIAL_MINERALS. Column A (the day number) is skipped.
    '''
    official = {_normalize_name(m['name']) for m in OFFICIAL_MINERALS}
    columns: Dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if idx == 0 or cell is None:
            continue
        raw = str(cell).split('\n')[0].strip()
        norm = _normalize_name(raw)
        if norm in official:
            columns[idx] = norm
    return columns


def _iter_daily_rows(sheet) -> List[Tuple[date, Dict[str, float]]]:
    '''
    Yields (price_date, {normalized_mineral: price}) for every numeric day row
    in a Diario sheet. Returns [] if the sheet does not match the expected
    layout (e.g. it is a Regalías sheet instead).
    '''
    rows = list(sheet.iter_rows(values_only = True))
    if len(rows) < 4:
        return []

    period = _parse_period_title(rows[1][0] if rows[1] else None)
    if period is None:
        return []
    year, month = period

    mineral_cols = _extract_mineral_columns(rows[2])
    if not mineral_cols:
        return []

    parsed: List[Tuple[date, Dict[str, float]]] = []
    for row in rows[3:]:
        day = row[0]
        if not isinstance(day, int):
            continue
        try:
            price_date = date(year, month, day)
        except ValueError:
            continue
        day_prices: Dict[str, float] = {}
        for col_idx, mineral_norm in mineral_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or val == '':
                continue
            try:
                day_prices[mineral_norm] = float(val)
            except (TypeError, ValueError):
                continue
        if day_prices:
            parsed.append((price_date, day_prices))
    return parsed


def _upsert_prices(
    session,
    mineral_id_by_norm: Dict[str, int],
    parsed_rows: List[Tuple[date, Dict[str, float]]]
) -> Tuple[int, int]:
    '''
    Inserts new MiningPrice rows for each (date, mineral) pair. Existing rows
    are left untouched so the operation is idempotent.

    Returns (inserted, skipped).
    '''
    inserted, skipped = 0, 0
    for price_date, prices in parsed_rows:
        for mineral_norm, value in prices.items():
            mineral_id = mineral_id_by_norm.get(mineral_norm)
            if mineral_id is None:
                skipped += 1
                continue
            exists = (
                session.query(MiningPrice.id)
                       .filter(MiningPrice.mineral_id == mineral_id,
                               MiningPrice.date == price_date)
                       .first()
            )
            if exists:
                skipped += 1
                continue
            session.add(MiningPrice(
                mineral_id = mineral_id,
                date = price_date,
                price_low = value,
                price_high = value,
            ))
            inserted += 1
    session.commit()
    return inserted, skipped


def main() -> int:
    '''
    CLI entry point.
    '''
    parser = argparse.ArgumentParser(
        description = 'Ingest biweekly Diario sheets from a curated xlsx.'
    )
    parser.add_argument('--source', required = True, type = Path,
                        help = 'Path to cotizaciones_mineras_bolivia.xlsx '
                               '(or similar layout).')
    parser.add_argument('--yes', action = 'store_true',
                        help = 'Apply the changes. Without this flag the '
                               'script only prints a plan.')
    args = parser.parse_args()

    if not args.source.exists():
        message = f'Source file not found: {args.source}'
        logger.error(message)
        print(message, file = sys.stderr)
        return 2

    workbook = openpyxl.load_workbook(args.source, data_only = True)
    diary_sheets = [s for s in workbook.sheetnames if 'diario' in s.lower()]
    if not diary_sheets:
        print('No Diario sheets detected. Nothing to do.')
        return 0

    all_parsed: List[Tuple[date, Dict[str, float]]] = []
    for sheet_name in diary_sheets:
        rows = _iter_daily_rows(workbook[sheet_name])
        print(f'[{sheet_name}] {len(rows)} day-rows with prices.')
        all_parsed.extend(rows)

    distinct_days = {d for d, _ in all_parsed}
    print(f'TOTAL: {len(distinct_days)} distinct dates across '
          f'{len(diary_sheets)} sheets.')

    if not args.yes:
        print('DRY-RUN: re-run with --yes to write to the database.')
        return 0

    session_factory = get_db_session(ENGINE)
    db_gen = session_factory()
    session = next(db_gen)
    try:
        ensure_official_minerals(session)
        mineral_id_by_norm = {
            _normalize_name(r.name): r.id for r in session.query(Mineral).all()
        }
        inserted, skipped = _upsert_prices(session, mineral_id_by_norm, all_parsed)
        message = f'Ingest complete: inserted={inserted}, skipped={skipped}.'
        logger.info(message)
        print(message)
        return 0
    finally:
        try:
            next(db_gen)
        except StopIteration:
            pass


if __name__ == '__main__':
    raise SystemExit(main())
