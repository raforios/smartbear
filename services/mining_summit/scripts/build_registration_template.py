'''
Generates the participant registration Excel template for the Cumbre Minera.

The template carries three sheets:
  - 'Registro'     : the data-entry grid (one participant per row).
  - 'Instrucciones': field-by-field rules (all fields but email are required).
  - 'Ejes Temáticos': the reference table of the six axes (values 1-6).

The EJE TEMÁTICO column is constrained to integers 1-6 via data validation
so the later ETL only has to confirm presence, not range.
'''
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Brand palette taken from the Cumbre Minera front-end config (oro / grafito).
GOLD = 'C9A751'
GRAPHITE = '242732'
GOLD_SOFT = 'F1E9CF'
WHITE = 'FFFFFF'

THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (header, key, required, help). Order matches the registration form.
FIELDS = [
    ('Carnet de Identidad', 'ci', True, 'Solo números. Identificador único del participante.'),
    ('Nombre', 'first_name', True, 'Nombre(s) del participante.'),
    ('Apellido', 'last_name', True, 'Apellido(s) del participante.'),
    ('Institución / Organización', 'institution', True,
     'Nombre de la institución a la que representa.'),
    ('Correo electrónico', 'email', False, 'ÚNICO campo opcional. Puede quedar vacío.'),
    ('Celular', 'phone', True, 'Número de contacto.'),
    ('Departamento', 'department', True,
     'La Paz, Cochabamba, Santa Cruz, Oruro, Potosí, Chuquisaca, Tarija, Beni o Pando.'),
    ('Eje Temático', 'axis', True, 'Número del 1 al 6 (ver hoja "Ejes Temáticos").'),
]

# Departments — mirror of config.json "departments" in the front-end.
DEPARTMENTS = [
    'La Paz', 'Cochabamba', 'Santa Cruz', 'Oruro', 'Potosí',
    'Chuquisaca', 'Tarija', 'Beni', 'Pando',
]

# (number, label) — mirror of summit_rules.AXIS_ORDER in the backend.
AXES = [
    (1, 'Seguridad Jurídica Minera'),
    (2, 'Contratos Mineros'),
    (3, 'Institucionalidad y Organización del Sector Minero'),
    (4, 'Medio Ambiente'),
    (5, 'Comercialización, Trazabilidad y Minería Ilegal'),
    (6, 'Desarrollo Productivo, Inversiones e Incentivos'),
]


def _field_index(key):
    '''Returns the 1-based column index of a field by its key.'''
    for index, (_header, field_key, _required, _help) in enumerate(FIELDS, start=1):
        if field_key == key:
            return index
    raise KeyError(key)


def _header_font():
    return Font(name='Calibri', size=11, bold=True, color=WHITE)


def _title_font():
    return Font(name='Calibri', size=15, bold=True, color=GRAPHITE)


def build_registro(worksheet):
    '''Builds the data-entry grid with a styled header row.'''
    for col_index, (header, _key, required, _help) in enumerate(FIELDS, start=1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.value = f'{header} *' if required else header
        cell.font = _header_font()
        cell.fill = PatternFill('solid', fgColor=GRAPHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        worksheet.column_dimensions[get_column_letter(col_index)].width = 26

    worksheet.row_dimensions[1].height = 30
    worksheet.freeze_panes = 'A2'

    # Restrict EJE TEMÁTICO (last column) to integers 1-6 for all entry rows.
    axis_col = get_column_letter(len(FIELDS))
    validation = DataValidation(
        type='whole', operator='between', formula1='1', formula2='6',
        allow_blank=True, showErrorMessage=True
    )
    validation.errorTitle = 'Eje temático inválido'
    validation.error = 'Ingrese un número entero del 1 al 6.'
    validation.promptTitle = 'Eje temático'
    validation.prompt = 'Elija un número del 1 al 6 (ver hoja "Ejes Temáticos").'
    worksheet.add_data_validation(validation)
    validation.add(f'{axis_col}2:{axis_col}1000')

    # Turn DEPARTAMENTO into a closed dropdown so registrants pick, not type.
    dept_col = get_column_letter(_field_index('department'))
    dept_validation = DataValidation(
        type='list', formula1=f'"{",".join(DEPARTMENTS)}"',
        allow_blank=True, showErrorMessage=True
    )
    dept_validation.errorTitle = 'Departamento inválido'
    dept_validation.error = 'Seleccione un departamento de la lista.'
    dept_validation.promptTitle = 'Departamento'
    dept_validation.prompt = 'Elija un departamento de la lista.'
    worksheet.add_data_validation(dept_validation)
    dept_validation.add(f'{dept_col}2:{dept_col}1000')


def build_instrucciones(worksheet):
    '''Builds the human-readable field rules sheet.'''
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 14
    worksheet.column_dimensions['C'].width = 60

    worksheet['A1'] = 'Instrucciones de llenado — Registro Cumbre Minera 2026'
    worksheet['A1'].font = _title_font()
    worksheet.merge_cells('A1:C1')
    worksheet.row_dimensions[1].height = 26

    note = worksheet['A2']
    note.value = ('Todos los campos son OBLIGATORIOS excepto "Correo electrónico". '
                  'Complete un participante por fila en la hoja "Registro".')
    note.font = Font(italic=True, color=GRAPHITE)
    worksheet.merge_cells('A2:C2')
    note.alignment = Alignment(wrap_text=True, vertical='center')
    worksheet.row_dimensions[2].height = 30

    headers = ['Campo', 'Obligatorio', 'Descripción']
    for col_index, text in enumerate(headers, start=1):
        cell = worksheet.cell(row=4, column=col_index)
        cell.value = text
        cell.font = _header_font()
        cell.fill = PatternFill('solid', fgColor=GRAPHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    for offset, (header, _key, required, help_text) in enumerate(FIELDS):
        row = 5 + offset
        values = [header, 'Sí' if required else 'No (opcional)', help_text]
        for col_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=col_index)
            cell.value = value
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if col_index == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=required, color='B82227' if not required else GRAPHITE)
            if offset % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F5F5F5')


def build_ejes(worksheet):
    '''Builds the reference table of the six thematic axes.'''
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 60

    worksheet['A1'] = 'Ejes Temáticos — Cumbre Minera 2026'
    worksheet['A1'].font = _title_font()
    worksheet.merge_cells('A1:B1')
    worksheet.row_dimensions[1].height = 26

    hint = worksheet['A2']
    hint.value = 'Escoja el eje de su preferencia e ingrese su número (1-6) en la hoja "Registro".'
    hint.font = Font(italic=True, color=GRAPHITE)
    worksheet.merge_cells('A2:B2')
    worksheet.row_dimensions[2].height = 20

    for col_index, text in enumerate(['N°', 'Eje Temático'], start=1):
        cell = worksheet.cell(row=4, column=col_index)
        cell.value = text
        cell.font = _header_font()
        cell.fill = PatternFill('solid', fgColor=GRAPHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

    for offset, (number, label) in enumerate(AXES):
        row = 5 + offset
        num_cell = worksheet.cell(row=row, column=1, value=number)
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.font = Font(bold=True, color=GRAPHITE)
        num_cell.fill = PatternFill('solid', fgColor=GOLD)
        num_cell.border = BORDER

        label_cell = worksheet.cell(row=row, column=2, value=label)
        label_cell.alignment = Alignment(wrap_text=True, vertical='center')
        label_cell.border = BORDER
        label_cell.fill = PatternFill('solid', fgColor=GOLD_SOFT)
        worksheet.row_dimensions[row].height = 22


def main(output_path):
    '''Builds the three-sheet registration template and writes it to disk.'''
    workbook = Workbook()
    registro = workbook.active
    registro.title = 'Registro'
    build_registro(registro)
    build_instrucciones(workbook.create_sheet('Instrucciones'))
    build_ejes(workbook.create_sheet('Ejes Temáticos'))
    workbook.save(output_path)
    print(f'Template written to {output_path}')


if __name__ == '__main__':
    import sys
    main(sys.argv[1])
