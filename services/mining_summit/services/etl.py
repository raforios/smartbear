'''
    ETL business logic: batch load of summit participants from an institution's
    Excel file.

    The institution is passed as a parameter (not read from the sheet), so every
    row in the file belongs to the same institution. The load validates the
    mandatory fields, enforces the institution cupo (only the first rows that fit
    are accredited), resolves each participant's chosen axis into a stable aula
    seat, and records the outcome as a load batch for evidence and follow-up.
'''
import io
import re
import unicodedata
import uuid
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from boto3.resources.base import ServiceResource
from fastapi import HTTPException

from services.crud import create_item
from services.environment import load_and_validate_env_vars
from services.exceptions import InvalidInputError
from services.institutions import get_institution
from services.logger_config import custom_logger as logger
from services.participants import count_active_by_institution, create_participant
from services.summit_rules import AXIS_BY_NUMBER
from services.utils import get_current_time_gmt, handle_service_errors

ENV_VARS = load_and_validate_env_vars({
    'DYNAMODB_TABLE_NAME_LOAD_BATCHES': str
})

LOAD_BATCHES_TABLE = ENV_VARS['DYNAMODB_TABLE_NAME_LOAD_BATCHES']

# Accepted header names per logical field (normalized: lowercase, no accents).
COLUMN_ALIASES: Dict[str, set] = {
    'ci': {'ci', 'carnet', 'carnet_de_identidad', 'documento', 'cedula'},
    'first_name': {'first_name', 'nombre', 'nombres'},
    'last_name': {'last_name', 'apellido', 'apellidos'},
    'email': {'email', 'correo', 'correo_electronico', 'e_mail', 'mail'},
    'phone': {'phone', 'celular', 'telefono', 'movil', 'tel'},
    'department': {'department', 'departamento', 'depto'},
    'axis': {'axis', 'eje', 'eje_tematico', 'eje_de_preferencia'}
}

# Fields required for a row to be accredited (email is the only optional one).
REQUIRED_FIELDS: Tuple[str, ...] = ('ci', 'first_name', 'last_name', 'phone', 'department', 'axis')


def _normalize(text: Any) -> str:
    '''Lowercases, strips accents and collapses non-alphanumerics to single spaces.'''
    ascii_text = ''.join(
        char for char in unicodedata.normalize('NFKD', str(text))
        if not unicodedata.combining(char)
    ).lower().strip()
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9]+', ' ', ascii_text)).strip()


def _header_field(header: Any) -> Optional[str]:
    '''Maps a spreadsheet header to a logical field name, or None.'''
    if header is None:
        return None
    key = _normalize(header).replace(' ', '_')
    for field, aliases in COLUMN_ALIASES.items():
        if key in aliases:
            return field
    return None


def _map_columns(header_row: Tuple[Any, ...]) -> Dict[int, str]:
    '''Builds a {column_index: field_name} map from the header row.'''
    mapping: Dict[int, str] = {}
    for index, header in enumerate(header_row):
        field = _header_field(header)
        if field:
            mapping[index] = field
    return mapping


def _parse_axis(raw: Any) -> Optional[str]:
    '''
        Resolves the raw "Eje Temático" cell (a 1-6 number) into an axis value.

        Returns:
            Optional[str]: The thematic axis value, or None if not a valid 1-6.
    '''
    if raw is None or str(raw).strip() == '':
        return None
    try:
        number = int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return None
    axis = AXIS_BY_NUMBER.get(number)
    return axis.value if axis else None


def _extract_rows(file_bytes: bytes) -> List[Tuple[int, Dict[str, Any]]]:
    '''
        Reads the workbook and returns (row_number, field_dict) tuples for every
        non-empty data row, using the header row to map columns to fields.

        Raises:
            InvalidInputError: If the file cannot be read as an Excel workbook.
    '''
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only = True)
    except Exception as error:
        raise InvalidInputError(
            detail = 'The uploaded file is not a valid Excel workbook.'
        ) from error

    worksheet = workbook['Registro'] if 'Registro' in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only = True))
    if not rows:
        return []

    column_map = _map_columns(rows[0])
    extracted: List[Tuple[int, Dict[str, Any]]] = []
    for offset, row in enumerate(rows[1:], start = 2):
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        fields = {
            field: row[index]
            for index, field in column_map.items()
            if index < len(row)
        }
        extracted.append((offset, fields))
    return extracted


def _validate_row(fields: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    '''
        Validates a raw row and returns a (clean_payload, error) tuple. Exactly
        one of the two is not None.
    '''
    clean: Dict[str, Any] = {}
    for field in ('ci', 'first_name', 'last_name', 'email', 'phone', 'department'):
        value = fields.get(field)
        clean[field] = str(value).strip() if value is not None else None

    axis = _parse_axis(fields.get('axis'))
    clean['axis'] = axis

    missing = [
        field for field in REQUIRED_FIELDS
        if not clean.get(field)
    ]
    if missing:
        if 'axis' in missing and fields.get('axis') not in (None, ''):
            return None, 'Invalid axis: must be a number from 1 to 6.'
        return None, f'Missing required fields: {", ".join(missing)}.'
    return clean, None


def _persist_batch(
    dynamodb_resource: ServiceResource,
    summary: Dict[str, Any]
) -> None:
    '''Stores the load-batch record (responsible + outcome) for evidence.'''
    create_item(
        dynamodb_resource = dynamodb_resource,
        table_name = LOAD_BATCHES_TABLE,
        item_data = {
            'batch_id': summary['batch_id'],
            'institution_id': summary['institution_id'],
            'institution_name': summary['institution_name'],
            'role': summary['role'],
            'responsible_name': summary['responsible']['name'],
            'responsible_phone': summary['responsible']['phone'],
            'cupos': summary['cupos'],
            'already_accredited': summary['already_accredited'],
            'accepted_count': summary['accepted_count'],
            'rejected_count': summary['rejected_count'],
            'accepted_cis': [entry['ci'] for entry in summary['accepted']],
            'rejected': summary['rejected'],
            'created_at': get_current_time_gmt().isoformat()
        },
        unique_key_attribute = 'batch_id'
    )


# pylint: disable=too-many-locals
@handle_service_errors
def load_participants_from_excel(
    dynamodb_resource: ServiceResource,
    file_bytes: bytes,
    institution_id: str,
    responsible: Dict[str, str],
    role: str
) -> Dict[str, Any]:
    '''
        Loads participants from an institution's Excel file, enforcing the
        institution cupo and assigning each accepted participant a stable
        eje/mesa seat. Every participant in the file is stamped with the given
        role. Records the outcome as a load batch.

        Args:
            dynamodb_resource (ServiceResource): The DynamoDB resource.
            file_bytes (bytes): The raw .xlsx content.
            institution_id (str): Institution the whole file belongs to.
            responsible (Dict[str, str]): {'name', 'phone'} of the responsible.
            role (str): Participant role assigned to every row of the file.

        Returns:
            Dict[str, Any]: The load summary (accepted + not-accredited report).
    '''
    institution = get_institution(dynamodb_resource, institution_id)
    cupos = int(institution['cupos'])
    already = count_active_by_institution(dynamodb_resource, institution_id)
    remaining = max(0, cupos - already)

    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []

    for row_number, fields in _extract_rows(file_bytes):
        clean, error = _validate_row(fields)
        if error:
            rejected.append({'row': row_number, 'ci': _safe_ci(fields), 'reason': error})
            continue
        if len(accepted) >= remaining:
            rejected.append({
                'row': row_number, 'ci': clean['ci'],
                'reason': f'Institution cupo reached ({cupos}); row not accredited.'
            })
            continue
        try:
            saved = create_participant(
                dynamodb_resource = dynamodb_resource,
                payload = {**clean, 'institution_id': institution_id, 'role': role}
            )
            accepted.append({
                'ci': saved['ci'],
                'first_name': saved['first_name'],
                'last_name': saved['last_name'],
                'axis': saved['axis'],
                'axis_label': saved['axis_label'],
                'mesa_code': saved['mesa_code']
            })
        except HTTPException as error:
            rejected.append({'row': row_number, 'ci': clean['ci'], 'reason': error.detail})

    summary = {
        'batch_id': str(uuid.uuid4()),
        'institution_id': institution_id,
        'institution_name': institution['name'],
        'role': role,
        'cupos': cupos,
        'already_accredited': already,
        'accepted_count': len(accepted),
        'rejected_count': len(rejected),
        'responsible': responsible,
        'accepted': accepted,
        'rejected': rejected
    }
    _persist_batch(dynamodb_resource, summary)
    message = (
        f'ETL load for institution={institution_id}: '
        f'{len(accepted)} accepted, {len(rejected)} rejected.'
    )
    logger.info(message)
    return summary


def _safe_ci(fields: Dict[str, Any]) -> Optional[str]:
    '''Returns a trimmed CI from a raw row, or None.'''
    value = fields.get('ci')
    return str(value).strip() if value is not None else None
