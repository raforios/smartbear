'''
    Excel export business logic for the Mining Summit reports.

    Builds .xlsx workbooks (as bytes) for the participants and attendances
    reports so the REPORTS/ADMIN roles can download them from the frontend.
'''
import io
from typing import Any, Dict, List, Sequence

import openpyxl
from boto3.resources.base import ServiceResource
from openpyxl.styles import Font, PatternFill

from services.attendances import ATTENDANCES_TABLE
from services.crud import scan_all_items
from services.participants import list_participants
from services.utils import handle_service_errors

_HEADER_FILL = PatternFill('solid', fgColor = '242732')
_HEADER_FONT = Font(bold = True, color = 'FFFFFF')

# (header, participant field) columns for the participants report.
_PARTICIPANT_COLUMNS: Sequence[tuple] = (
    ('CI', 'ci'),
    ('Nombre', 'first_name'),
    ('Apellido', 'last_name'),
    ('Institución', 'institution_name'),
    ('Rol', 'role'),
    ('Eje Temático', 'axis_label'),
    ('Aula', 'mesa_code'),
    ('Departamento', 'department'),
    ('Correo', 'email'),
    ('Celular', 'phone'),
    ('Estado', 'status'),
    ('Observación', 'observation'),
    ('Registrado', 'registered_at')
)

# (header, attendance field) columns for the attendances report.
_ATTENDANCE_COLUMNS: Sequence[tuple] = (
    ('CI', 'ci'),
    ('Fecha', 'attendance_date'),
    ('Hora', 'attendance_at'),
    ('Registrado por', 'marked_by')
)


def _write_workbook(
    title: str,
    columns: Sequence[tuple],
    rows: List[Dict[str, Any]]
) -> bytes:
    '''
        Builds an .xlsx workbook with a styled header and returns its bytes.

        Args:
            title (str): Worksheet title.
            columns (Sequence[tuple]): (header, field) column definitions.
            rows (List[Dict[str, Any]]): Records to dump.

        Returns:
            bytes: The serialized workbook.
    '''
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title

    for col_index, (header, _field) in enumerate(columns, start = 1):
        cell = worksheet.cell(row = 1, column = col_index, value = header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        worksheet.column_dimensions[cell.column_letter].width = 20
    worksheet.freeze_panes = 'A2'

    for row_index, record in enumerate(rows, start = 2):
        for col_index, (_header, field) in enumerate(columns, start = 1):
            worksheet.cell(row = row_index, column = col_index, value = record.get(field))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@handle_service_errors
def export_participants_xlsx(
    dynamodb_resource: ServiceResource,
    include_inactive: bool = False
) -> bytes:
    '''
        Builds the participants report workbook. By default only ACTIVE
        participants are exported (the initial report); pass include_inactive to
        include replaced/cancelled ones too.
    '''
    response = list_participants(
        dynamodb_resource = dynamodb_resource,
        query_params = {'include_inactive': include_inactive, 'limit': 100000}
    )
    items = response['items']
    items.sort(key = lambda item: (item.get('institution_name') or '', item.get('last_name') or ''))
    return _write_workbook('Participantes', _PARTICIPANT_COLUMNS, items)


@handle_service_errors
def export_attendances_xlsx(dynamodb_resource: ServiceResource) -> bytes:
    '''
        Builds the attendances report workbook covering every recorded check-in.
    '''
    items = scan_all_items(dynamodb_resource, ATTENDANCES_TABLE)
    items.sort(key = lambda item: (item.get('attendance_date') or '', item.get('ci') or ''))
    return _write_workbook('Asistencias', _ATTENDANCE_COLUMNS, items)
